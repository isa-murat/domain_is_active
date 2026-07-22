import datetime
import os
from typing import List, Dict, Any
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from domain_is_active.constants.enums import ReportColors, ScanDecision


class ExcelExporter:
    """
    Tarama sonuçlarını profesyonel Yönetici Özeti ve Detaylı Analiz
    sayfaları içeren tıklanabilir Excel raporuna dönüştüren modül.
    """

    def __init__(self, results: List[Dict[str, Any]]):
        self.results = results

    def export(self, output_path: str = None, silent: bool = False) -> str:
        """
        Excel raporunu üretir ve belirtilen dosya yoluna kaydeder.
        """
        if not output_path:
            timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join("reports", f"phishing_analysis_report_{timestamp_str}.xlsx")

        abs_output_path = os.path.abspath(output_path)
        os.makedirs(os.path.dirname(abs_output_path), exist_ok=True)

        if not silent:
            print(f"[*] Excel raporu oluşturuluyor: {abs_output_path}")
        df = pd.DataFrame(self.results)

        wb = openpyxl.Workbook()

        # SAYFA 1: Yönetici Özeti
        ws_summary = wb.active
        ws_summary.title = "Yonetici Ozet"
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary.merge_cells("A1:D1")
        title_cell = ws_summary["A1"]
        title_cell.value = "PHISHING DOMAIN AKTIFLIK VE TEHDIT AVCILIGI OZET RAPORU"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=ReportColors.TITLE_BG, end_color=ReportColors.TITLE_BG, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.row_dimensions[1].height = 35

        total_count = len(df)
        decision_counts = df["decision"].value_counts().to_dict() if not df.empty and "decision" in df.columns else {}

        ws_summary["A3"] = "Metrik / Durum"
        ws_summary["B3"] = "Domain Sayısı"
        ws_summary["C3"] = "Oran (%)"

        for col in ["A3", "B3", "C3"]:
            cell = ws_summary[col]
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=ReportColors.HEADER_BG, end_color=ReportColors.HEADER_BG, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        stats_rows = [
            ("Toplam Taranan Domain", total_count),
            ("ACTIVE (Aktif / Canlı)", decision_counts.get(ScanDecision.ACTIVE, 0)),
            ("TAKEDOWN (Kapatılmış)", decision_counts.get(ScanDecision.TAKEDOWN, 0)),
            ("INACTIVE (Pasif / Ölü)", decision_counts.get(ScanDecision.INACTIVE, 0)),
            ("SUSPICIOUS / UNSTABLE (Şüpheli)", decision_counts.get(ScanDecision.SUSPICIOUS, 0)),
            ("PARKED (Park Edilmiş)", decision_counts.get(ScanDecision.PARKED, 0)),
        ]

        for r_idx, (label, val) in enumerate(stats_rows, start=4):
            ws_summary.cell(row=r_idx, column=1, value=label).font = Font(bold=(r_idx == 4))
            ws_summary.cell(row=r_idx, column=2, value=val).alignment = Alignment(horizontal="center")
            pct = (val / total_count * 100) if total_count > 0 else 0
            ws_summary.cell(row=r_idx, column=3, value=f"{pct:.1f}%").alignment = Alignment(horizontal="center")

        # SAYFA 2: Detaylı Analiz
        ws_detail = wb.create_sheet(title="Detayli Analiz")
        ws_detail.views.sheetView[0].showGridLines = True

        headers = [
            "Domain", "Nihai Karar", "Karar Gerekçesi", "DNS Status",
            "IPv4 Adresleri", "IPv6 Adresleri", "HTTP Status", "Yönlendirme URL", "SSL Geçerli",
            "SSL Issuer", "Favicon SHA256", "SPKI SHA256", "WHOIS Hold",
            "URLScan Geçmiş", "URLScan Tarih", "Ekran Görüntüsü", "İlişkili Avlanan Domainler"
        ]

        ws_detail.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col_num)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=ReportColors.TITLE_BG, end_color=ReportColors.TITLE_BG, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_detail.row_dimensions[1].height = 28

        domain_to_row = {}
        for idx, item in enumerate(self.results, start=2):
            domain_to_row[str(item.get("domain", "")).lower()] = idx

        color_map = {
            ScanDecision.ACTIVE: PatternFill(start_color=ReportColors.ACTIVE_BG, end_color=ReportColors.ACTIVE_BG, fill_type="solid"),
            ScanDecision.TAKEDOWN: PatternFill(start_color=ReportColors.TAKEDOWN_BG, end_color=ReportColors.TAKEDOWN_BG, fill_type="solid"),
            ScanDecision.INACTIVE: PatternFill(start_color=ReportColors.INACTIVE_BG, end_color=ReportColors.INACTIVE_BG, fill_type="solid"),
            ScanDecision.SUSPICIOUS: PatternFill(start_color=ReportColors.SUSPICIOUS_BG, end_color=ReportColors.SUSPICIOUS_BG, fill_type="solid"),
            ScanDecision.PARKED: PatternFill(start_color=ReportColors.PARKED_BG, end_color=ReportColors.PARKED_BG, fill_type="solid"),
        }

        for row_idx, item in enumerate(self.results, start=2):
            row_data = [
                item.get("domain", ""),
                str(item.get("decision", "")),
                item.get("reason", ""),
                item.get("dns_resolved", "Hayır"),
                item.get("ipv4_addresses", "-"),
                item.get("ipv6_addresses", "-"),
                item.get("http_status", "-"),
                item.get("redirect_url", "-"),
                item.get("ssl_valid", "Hayır"),
                item.get("ssl_issuer", "-"),
                item.get("favicon_sha256", "-"),
                item.get("spki_sha256", "-"),
                item.get("whois_hold", "Hayır"),
                item.get("urlscan_history", "Hayır"),
                item.get("urlscan_time", "-"),
                "",
                ""
            ]
            ws_detail.append(row_data)

            decision_val = item.get("decision")
            if decision_val in color_map:
                ws_detail.cell(row=row_idx, column=2).fill = color_map[decision_val]
                ws_detail.cell(row=row_idx, column=2).font = Font(bold=True)

            ss_url = item.get("screenshot_url")
            ss_cell = ws_detail.cell(row=row_idx, column=16)
            if ss_url and ss_url != "-":
                ss_cell.value = f'=HYPERLINK("{ss_url}", "Görseli Aç 🔗")'
                ss_cell.font = Font(color="0000FF", underline="single")
            else:
                ss_cell.value = "-"

            corr_str = item.get("correlated_domains")
            corr_cell = ws_detail.cell(row=row_idx, column=17)

            if corr_str and corr_str != "-":
                first_corr = str(corr_str).split(",")[0].strip().lower()
                if first_corr in domain_to_row:
                    target_row = domain_to_row[first_corr]
                    corr_cell.value = f'=HYPERLINK("#\'Detayli Analiz\'!A{target_row}", "{corr_str}")'
                    corr_cell.font = Font(color="0000FF", underline="single", bold=True)
                else:
                    corr_cell.value = corr_str
            else:
                corr_cell.value = "-"

        for col in ws_detail.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_detail.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        for col in ws_summary.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 18)

        wb.save(abs_output_path)
        if not silent:
            print(f"[+] Excel raporu başarıyla kaydedildi: {abs_output_path}")
        return abs_output_path
