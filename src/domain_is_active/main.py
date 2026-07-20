import argparse
import datetime
import os
import re
import sys
import time
from collections import deque
from concurrent.futures import ThreadPoolExecutor
from typing import Dict, Any

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Add src directory to sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

# Local module imports
from domain_is_active.checker import PhishingDomainChecker
from domain_is_active.hunter import URLScanHunter

# Suppress insecure request warnings
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def sanitize_domain(raw_domain: str) -> str:
    """Domain adını temizler (protokol, port veya path ifadelerini kaldırır)."""
    if not raw_domain or not isinstance(raw_domain, str):
        return ""
    d = raw_domain.strip().lower()
    # Protokol temizleme (http:// veya https://)
    d = re.sub(r"^https?://", "", d)
    # Path, query, port temizleme
    d = d.split("/")[0].split("?")[0].split("#")[0].split(":")[0].strip()
    return d


class PhishingPipelineOrchestrator:
    def __init__(
        self,
        input_path: str,
        max_threads: int = 10,
        max_correlated_per_domain: int = 3,
    ):
        self.input_path = input_path
        self.max_threads = max_threads
        self.max_correlated_per_domain = max_correlated_per_domain
        self.queue = deque()
        self.visited_domains = set()
        self.results = []
        self.hunter = URLScanHunter()

    def load_initial_domains(self):
        """
        Belirtilen girdi dosyasından (.csv, .txt, .xlsx) domainleri okur ve kuyruğa ekler.
        """
        if not os.path.exists(self.input_path):
            print(f"[!] Hata: Girdi dosyası bulunamadı: {self.input_path}")
            sys.exit(1)

        ext = os.path.splitext(self.input_path)[1].lower()
        raw_domains = []

        try:
            if ext in [".csv", ".tsv"]:
                df = pd.read_csv(self.input_path)
                # Domain içeren muhtemel sütun isimlerini arıyoruz
                target_col = None
                for col in df.columns:
                    if any(keyword in str(col).lower() for keyword in ["domain", "url", "site", "host"]):
                        target_col = col
                        break
                if target_col is None:
                    target_col = df.columns[0]
                raw_domains = df[target_col].dropna().tolist()

            elif ext in [".xlsx", ".xls"]:
                df = pd.read_excel(self.input_path)
                target_col = None
                for col in df.columns:
                    if any(keyword in str(col).lower() for keyword in ["domain", "url", "site", "host"]):
                        target_col = col
                        break
                if target_col is None:
                    target_col = df.columns[0]
                raw_domains = df[target_col].dropna().tolist()

            else:
                # Metin dosyası varsayılanı (satır satır okuma)
                with open(self.input_path, "r", encoding="utf-8", errors="ignore") as f:
                    raw_domains = [line.strip() for line in f if line.strip()]

        except Exception as e:
            print(f"[!] Girdi dosyası okunurken hata oluştu: {e}")
            sys.exit(1)

        for raw in raw_domains:
            domain_clean = sanitize_domain(str(raw))
            if domain_clean and domain_clean not in self.visited_domains:
                self.visited_domains.add(domain_clean)
                self.queue.append(domain_clean)

        print(f"[*] Girdi dosyasından ({self.input_path}) {len(self.queue)} adet başlangıç domaini kuyruğa yüklendi.")

    def process_single_domain(self, domain: str) -> Dict[str, Any]:
        """
        Bir domain için Pipeline adımlarını çalıştırır:
        1. Lokal Checker Analizi
        2. URLScan Geçmiş Sorgusu (Geçmiş & Ekran Görüntüsü)
        3. URLScan Parmak İzi Korelasyonu (Tehdit Avcılığı)
        """
        print(f"[->] Taranıyor: {domain}")

        # 1. Aşama: Lokal Analiz Motoru
        checker = PhishingDomainChecker(domain)
        local_res = checker.run()

        # 2. Aşama: URLScan Geçmiş Sorgusu
        history = self.hunter.get_historical_data(domain)

        # 3. Aşama: Tehdit Avcılığı (Favicon veya SPKI ile Tersine Arama)
        favicon_sha256 = local_res.get("favicon_sha256")
        spki_sha256 = local_res.get("spki_sha256")

        correlated_domains = []
        if favicon_sha256:
            correlated_domains = self.hunter.correlate_fingerprints(favicon_sha256=favicon_sha256)
        elif spki_sha256:
            correlated_domains = self.hunter.correlate_fingerprints(spki_hash=spki_sha256)

        # Kendi kendini ilişkili listeden çıkar
        correlated_domains = [d for d in correlated_domains if d.lower() != domain.lower()]

        # 4. Aşama: Dinamik Geri Besleme (Kuyruğa Yeni Domain Ekleme)
        added_count = 0
        for corr_domain in correlated_domains[: self.max_correlated_per_domain]:
            corr_clean = sanitize_domain(corr_domain)
            if corr_clean and corr_clean not in self.visited_domains:
                self.visited_domains.add(corr_clean)
                self.queue.append(corr_clean)
                added_count += 1

        if added_count > 0:
            print(f"  [+] Tehdit Avcılığı: {domain} üzerinden {added_count} yeni ilişkili domain kuyruğa eklendi.")

        record = {
            "domain": domain,
            "decision": local_res["decision"],
            "reason": local_res["reason"],
            "dns_resolved": "Evet" if local_res["dns_resolved"] else "Hayır",
            "ipv4_addresses": ", ".join(local_res["ipv4_addresses"]) or "-",
            "ipv6_addresses": ", ".join(local_res["ipv6_adressess"]) or "-",
            "http_status": str(local_res["http_status"]) if local_res["http_status"] else "-",
            "redirect_url": local_res["redirect_url"] or "-",
            "ssl_valid": "Evet" if local_res["ssl_valid"] else "Hayır",
            "ssl_issuer": local_res["ssl_issuer"] or "-",
            "favicon_sha256": favicon_sha256 or "-",
            "spki_sha256": spki_sha256 or "-",
            "whois_hold": "Evet" if local_res["whois_hold"] else "Hayır",
            "urlscan_history": "Evet" if history.get("has_history") else "Hayır",
            "urlscan_time": history.get("scan_time") or "-",
            "screenshot_url": history.get("screenshot_url") or "-",
            "correlated_domains": ", ".join(correlated_domains) if correlated_domains else "-",
        }
        return record

    def run_pipeline(self):
        """Kuyruktaki tüm domainler (başlangıç + korelasyon ile bulunanlar) bitene kadar parallel işçilerle çalışır."""
        self.load_initial_domains()

        start_time = time.time()
        print(f"\n[*] Pipeline Başlatıldı ({self.max_threads} eşzamanlı istek işçisi ile)...")

        while self.queue:
            batch = []
            while self.queue and len(batch) < self.max_threads:
                batch.append(self.queue.popleft())

            with ThreadPoolExecutor(max_workers=self.max_threads) as executor:
                batch_results = list(executor.map(self.process_single_domain, batch))
                self.results.extend(batch_results)

        elapsed = time.time() - start_time
        print(f"\n[+] Pipeline Tamamlandı! Toplam {len(self.results)} domain {elapsed:.2f} saniyede analiz edildi.")

    def export_excel_report(self, output_path: str = None) -> str:
        """
        Sonuçları Yönetici Özeti ve Detaylı Analiz sayfaları içeren tıklanabilir Excel raporuna dönüştürür.
        """
        if not output_path:
            timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join("reports", f"phishing_analysis_report_{timestamp_str}.xlsx")

        abs_output_path = os.path.abspath(output_path)
        os.makedirs(os.path.dirname(abs_output_path), exist_ok=True)

        print(f"[*] Excel raporu oluşturuluyor: {abs_output_path}")
        df = pd.DataFrame(self.results)

        wb = openpyxl.Workbook()

        # SAYFA 1: Yönetici Özeti
        ws_summary = wb.active
        ws_summary.title = "Yonetici Ozet"
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary.merge_cells("A1:D1")
        title_cell = ws_summary["A1"]
        title_cell.value = "PHISHING DOMAIN AKTIFLIK VE TEHDIT AVCILIGI OZET RAPORU"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.row_dimensions[1].height = 35

        total_count = len(df)
        decision_counts = df["decision"].value_counts().to_dict() if not df.empty else {}

        ws_summary["A3"] = "Metrik / Durum"
        ws_summary["B3"] = "Domain Sayısı"
        ws_summary["C3"] = "Oran (%)"

        for col in ["A3", "B3", "C3"]:
            cell = ws_summary[col]
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        stats_rows = [
            ("Toplam Taranan Domain", total_count),
            ("ACTIVE (Aktif / Canlı)", decision_counts.get("ACTIVE (AKTIF)", 0)),
            ("TAKEDOWN (Kapatılmış)", decision_counts.get("TAKEDOWN (KAPATILDI)", 0)),
            ("INACTIVE (Pasif / Ölü)", decision_counts.get("INACTIVE (PASIF)", 0)),
            ("SUSPICIOUS / UNSTABLE (Şüpheli)", decision_counts.get("SUSPICIOUS / UNSTABLE", 0)),
            ("PARKED (Park Edilmiş)", decision_counts.get("PARKED (PARK EDILMIS)", 0)),
        ]

        for r_idx, (label, val) in enumerate(stats_rows, start=4):
            ws_summary.cell(row=r_idx, column=1, value=label).font = Font(bold=(r_idx == 4))
            ws_summary.cell(row=r_idx, column=2, value=val).alignment = Alignment(horizontal="center")
            pct = (val / total_count * 100) if total_count > 0 else 0
            ws_summary.cell(row=r_idx, column=3, value=f"{pct:.1f}%").alignment = Alignment(horizontal="center")

        # SAYFA 2: Detaylı Analiz
        ws_detail = wb.create_sheet(title="Detayli Analiz")
        ws_detail.views.sheetView[0].showGridLines = True

        headers = [
            "Domain", "Nihai Karar", "Karar Gerekçesi", "DNS Status",
            "IPv4 Adresleri", "IPv6 Adresleri", "HTTP Status", "Yönlendirme URL", "SSL Geçerli",
            "SSL Issuer", "Favicon SHA256", "SPKI SHA256", "WHOIS Hold",
            "URLScan Geçmiş", "URLScan Tarih", "Ekran Görüntüsü", "İlişkili Avlanan Domainler"
        ]

        ws_detail.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col_num)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_detail.row_dimensions[1].height = 28

        domain_to_row = {}
        for idx, item in enumerate(self.results, start=2):
            domain_to_row[item["domain"].lower()] = idx

        color_map = {
            "ACTIVE (AKTIF)": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "TAKEDOWN (KAPATILDI)": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "INACTIVE (PASIF)": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "SUSPICIOUS / UNSTABLE": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "PARKED (PARK EDILMIS)": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        }

        for row_idx, item in enumerate(self.results, start=2):
            row_data = [
                item["domain"],
                item["decision"],
                item["reason"],
                item["dns_resolved"],
                item["ipv4_addresses"],
                item["ipv6_addresses"],
                item["http_status"],
                item["redirect_url"],
                item["ssl_valid"],
                item["ssl_issuer"],
                item["favicon_sha256"],
                item["spki_sha256"],
                item["whois_hold"],
                item["urlscan_history"],
                item["urlscan_time"],
                "",
                ""
            ]
            ws_detail.append(row_data)

            decision_val = item["decision"]
            if decision_val in color_map:
                ws_detail.cell(row=row_idx, column=2).fill = color_map[decision_val]
                ws_detail.cell(row=row_idx, column=2).font = Font(bold=True)

            ss_url = item["screenshot_url"]
            ss_cell = ws_detail.cell(row=row_idx, column=16)
            if ss_url and ss_url != "-":
                ss_cell.value = f'=HYPERLINK("{ss_url}", "Görseli Aç 🔗")'
                ss_cell.font = Font(color="0000FF", underline="single")
            else:
                ss_cell.value = "-"

            corr_str = item["correlated_domains"]
            corr_cell = ws_detail.cell(row=row_idx, column=17)

            if corr_str and corr_str != "-":
                first_corr = corr_str.split(",")[0].strip().lower()
                if first_corr in domain_to_row:
                    target_row = domain_to_row[first_corr]
                    corr_cell.value = f'=HYPERLINK("#\'Detayli Analiz\'!A{target_row}", "{corr_str}")'
                    corr_cell.font = Font(color="0000FF", underline="single", bold=True)
                else:
                    corr_cell.value = corr_str
            else:
                corr_cell.value = "-"

        for col in ws_detail.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_detail.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        for col in ws_summary.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 18)

        wb.save(abs_output_path)
        print(f"[+] Excel raporu başarıyla kaydedildi: {abs_output_path}")
        return abs_output_path


def main():
    parser = argparse.ArgumentParser(
        description="Phishing Active & Correlation Tool - Alan Adı Aktiflik ve Tehdit Avcılığı Analizi"
    )
    parser.add_argument(
        "-p", "--path",
        type=str,
        required=True,
        help="Girdi dosyası yolu (.csv, .txt, .xlsx)"
    )
    parser.add_argument(
        "-o", "--output",
        type=str,
        default=None,
        help="Rapor Excel çıktısının kaydedileceği yol (Varsayılan: reports/phishing_analysis_report_<timestamp>.xlsx)"
    )
    parser.add_argument(
        "-c", "--max-correlated",
        type=int,
        default=3,
        help="Domain başı kuyruğa eklenecek max ilişkili domain sayısı (Varsayılan: 3)"
    )

    args = parser.parse_args()

    orchestrator = PhishingPipelineOrchestrator(
        input_path=args.path,
        max_correlated_per_domain=args.max_correlated,
    )
    orchestrator.run_pipeline()
    orchestrator.export_excel_report(args.output)


if __name__ == "__main__":
    main()
